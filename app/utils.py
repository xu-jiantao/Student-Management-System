from __future__ import annotations

import os
from datetime import datetime
from functools import wraps
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Iterable

from flask import abort, current_app, flash, redirect, request, url_for
from flask_login import current_user
from itsdangerous import BadSignature, URLSafeTimedSerializer
from openpyxl import Workbook, load_workbook

from .extensions import db
from .models import OperationLog, Role


def get_upload_path() -> Path:
    upload_folder = Path(current_app.config.get("UPLOAD_FOLDER", "uploads"))
    upload_folder.mkdir(parents=True, exist_ok=True)
    return upload_folder


def save_uploaded_file(file_storage, subdir: str = "") -> str:
    upload_path = get_upload_path()
    if subdir:
        upload_path = upload_path / subdir
        upload_path.mkdir(parents=True, exist_ok=True)

    filename = file_storage.filename or "uploaded"
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    stored_name = f"{timestamp}_{filename}"
    stored_path = upload_path / stored_name
    file_storage.save(stored_path)
    static_folder = Path(current_app.static_folder)
    return str(stored_path.relative_to(static_folder))


def role_required(role_names: Iterable[str]) -> Callable:
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("auth.login", next=request.path))
            if not any(current_user.has_role(role) for role in role_names):
                abort(403)
            return func(*args, **kwargs)

        return wrapper

    return decorator


def permission_required(code: str) -> Callable:
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for("auth.login", next=request.path))
            if not current_user.has_permission(code):
                abort(403)
            return func(*args, **kwargs)

        return wrapper

    return decorator


def generate_token(user_id: int, purpose: str = "password-reset") -> str:
    serializer = URLSafeTimedSerializer(current_app.config["SECRET_KEY"])
    return serializer.dumps({"user_id": user_id, "purpose": purpose})


def verify_token(token: str, max_age: int = 3600, purpose: str = "password-reset") -> int | None:
    serializer = URLSafeTimedSerializer(current_app.config["SECRET_KEY"])
    try:
        data = serializer.loads(token, max_age=max_age)
    except BadSignature:
        return None
    if data.get("purpose") != purpose:
        return None
    return data.get("user_id")


def log_operation(user_id: int | None, action: str, resource: str, description: str = "") -> None:
    log = OperationLog(
        user_id=user_id,
        action=action,
        resource=resource,
        description=description,
        ip_address=request.remote_addr if request else None,
    )
    db.session.add(log)
    db.session.commit()


def paginate_query(query, page: int, per_page: int = 20):
    items = query.limit(per_page).offset((page - 1) * per_page).all()
    total = query.order_by(None).count()
    return items, total


def ensure_roles_exist(role_names: list[str]) -> None:
    for role_name in role_names:
        if not Role.query.filter_by(name=role_name).first():
            role = Role(name=role_name)
            db.session.add(role)
    db.session.commit()


def flash_form_errors(errors: dict[str, list[str]]) -> None:
    for messages in errors.values():
        for message in messages:
            flash(message, "danger")


def import_students_from_excel(file_path: str) -> list[dict[str, Any]]:
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    students: list[dict[str, Any]] = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    header_index = {header: index for index, header in enumerate(headers)}
    required_fields = {"student_number", "name", "gender", "date_of_birth", "class_name", "email", "phone"}
    missing_fields = required_fields - set(headers)
    if missing_fields:
        raise ValueError(f"缺少必填列: {', '.join(missing_fields)}")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        student = {
            "student_number": row[header_index["student_number"]],
            "name": row[header_index["name"]],
            "gender": row[header_index["gender"]],
            "date_of_birth": row[header_index["date_of_birth"]],
            "class_name": row[header_index["class_name"]],
            "email": row[header_index["email"]],
            "phone": row[header_index["phone"]],
            "address": row[header_index.get("address", -1)] if header_index.get("address") is not None else "",
        }
        students.append(student)
    return students


def export_students_to_excel(rows: Iterable[dict[str, Any]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    headers = [
        "student_number",
        "name",
        "gender",
        "date_of_birth",
        "class_name",
        "email",
        "phone",
        "address",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def export_grades_to_excel(rows: Iterable[dict[str, Any]]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    headers = ["student_number", "student_name", "course", "term", "assessment_type", "score", "recorded_at"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
